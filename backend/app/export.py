"""Export utilities for projects data.

支持导出格式:
- Excel (.xlsx)
- CSV (.csv)
"""

import csv
import io
from typing import Any

import pandas as pd
import structlog
from openpyxl.styles import Alignment, Font, PatternFill

logger = structlog.get_logger(__name__)


def export_projects_to_excel(projects: list[dict[str, Any]]) -> bytes:
    """导出项目列表到 Excel 格式.

    Args:
        projects: 项目列表

    Returns:
        Excel 文件的字节数据
    """
    # 准备数据
    data = []
    for project in projects:
        row = {
            "ID": project.get("id", ""),
            "项目名称": project.get("name", ""),
            "URL": project.get("url", ""),
            "赛道": project.get("sector", ""),
            "阶段": project.get("stage", ""),
            "评分": project.get("score", 0),
            "标签": project.get("label", ""),
            "置信度": project.get("confidence", 0.0),
            "创建时间": project.get("created_at", ""),
            "更新时间": project.get("updated_at", ""),
        }
        data.append(row)

    # 创建 DataFrame
    df = pd.DataFrame(data)

    # 写入 Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Projects", index=False)

        # 获取工作表
        worksheet = writer.sheets["Projects"]

        # 设置样式
        # 标题行样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 调整列宽
        column_widths = {
            "A": 15,  # ID
            "B": 20,  # 项目名称
            "C": 40,  # URL
            "D": 15,  # 赛道
            "E": 15,  # 阶段
            "F": 10,  # 评分
            "G": 10,  # 标签
            "H": 12,  # 置信度
            "I": 20,  # 创建时间
            "J": 20,  # 更新时间
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        # 设置数据行样式
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # 根据标签设置颜色
            label = row[6].value  # G列是标签
            if label == "FARM":
                fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif label == "WATCH":
                fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif label == "IGNORE":
                fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            else:
                fill = None

            if fill:
                for cell in row:
                    cell.fill = fill

    output.seek(0)
    return output.read()


def export_projects_to_csv(projects: list[dict[str, Any]]) -> str:
    """导出项目列表到 CSV 格式.

    Args:
        projects: 项目列表

    Returns:
        CSV 文件的字符串数据
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # 写入表头
    headers = ["ID", "项目名称", "URL", "赛道", "阶段", "评分", "标签", "置信度", "创建时间", "更新时间"]
    writer.writerow(headers)

    # 写入数据
    for project in projects:
        row = [
            project.get("id", ""),
            project.get("name", ""),
            project.get("url", ""),
            project.get("sector", ""),
            project.get("stage", ""),
            project.get("score", 0),
            project.get("label", ""),
            project.get("confidence", 0.0),
            project.get("created_at", ""),
            project.get("updated_at", ""),
        ]
        writer.writerow(row)

    return output.getvalue()


def export_project_detail_to_excel(project: dict[str, Any]) -> bytes:
    """导出单个项目详情到 Excel（包含完整分析结果）.

    Args:
        project: 项目详情

    Returns:
        Excel 文件的字节数据
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Sheet 1: 基本信息
        basic_data = {
            "字段": ["ID", "项目名称", "URL", "赛道", "阶段", "评分", "标签", "置信度"],
            "值": [
                project.get("id", ""),
                project.get("name", ""),
                project.get("url", ""),
                project.get("sector", ""),
                project.get("stage", ""),
                project.get("score", 0),
                project.get("label", ""),
                project.get("confidence", 0.0),
            ],
        }
        df_basic = pd.DataFrame(basic_data)
        df_basic.to_excel(writer, sheet_name="基本信息", index=False)

        # Sheet 2: 评分原因
        if project.get("reason"):
            reason_data = {"原因": project["reason"]}
            df_reason = pd.DataFrame(reason_data)
            df_reason.to_excel(writer, sheet_name="评分原因", index=False)

        # Sheet 3: 叙事分析
        if project.get("narrative"):
            narrative = project["narrative"]
            narrative_data = {
                "字段": ["赛道", "阶段", "热度分数", "时机"],
                "值": [
                    narrative.get("sector", ""),
                    narrative.get("stage", ""),
                    narrative.get("heat_score", 0),
                    narrative.get("timing", ""),
                ],
            }
            df_narrative = pd.DataFrame(narrative_data)
            df_narrative.to_excel(writer, sheet_name="叙事分析", index=False)

        # Sheet 4: 团队分析
        if project.get("team"):
            team = project["team"]
            team_data = {
                "字段": ["团队分数", "团队类型", "团队标签"],
                "值": [
                    team.get("team_score", 0),
                    team.get("team_type", ""),
                    ", ".join(team.get("team_flags", [])),
                ],
            }
            df_team = pd.DataFrame(team_data)
            df_team.to_excel(writer, sheet_name="团队分析", index=False)

        # Sheet 5: 风险分析
        if project.get("risk"):
            risk = project["risk"]
            risk_data = {
                "字段": ["代币风险", "解锁压力", "风险标签"],
                "值": [
                    risk.get("token_risk", 0),
                    risk.get("unlock_pressure", ""),
                    ", ".join(risk.get("risk_flags", [])),
                ],
            }
            df_risk = pd.DataFrame(risk_data)
            df_risk.to_excel(writer, sheet_name="风险分析", index=False)

        # Sheet 6: 代币经济学
        if project.get("tokenomics"):
            tokenomics = project["tokenomics"]
            tokenomics_data = {
                "字段": ["VC 占比", "团队占比", "解锁惩罚"],
                "值": [
                    tokenomics.get("vc_share", 0),
                    tokenomics.get("team_share", 0),
                    tokenomics.get("unlock_penalty", 0),
                ],
            }
            df_tokenomics = pd.DataFrame(tokenomics_data)
            df_tokenomics.to_excel(writer, sheet_name="代币经济学", index=False)

        # 美化所有 sheet
        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            # 标题行样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 调整列宽
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception as exc:
                        logger.debug("export.column_width_failed", error=str(exc))
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

    output.seek(0)
    return output.read()
